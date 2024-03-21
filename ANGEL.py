from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as ec
import openpyxl
from selenium.webdriver.support.wait import WebDriverWait
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options
from datetime import datetime

current_time = datetime.now()
formatted_time = current_time.strftime("%a, %d %b %Y %H:%M:%S")
print(f"INFO  @ {formatted_time}:\n\tRunning ANGEL...\n")
print('''
ANGEL (version 2.0)
Automation NGS primer design for Gene Editing in Large scale. Specifically tailored for genome engineering 
applications, such as nuclease editing, base editing, and prime editing. The tool utilizes the UCSC Genome 
Browser to identify flanking genomic sequences of the target. NCBI Primer Blast is used to pinpoint primers 
that are specific to the target region. The UCSC in silico PCR tool is used to verify that the primer pair 
identified by NCBI Primer Blast are specific and capable of producing only the intended amplicon sequence. 

                         .=+#%+.                                                                    
                     ..  .-%%%#:                                                                    
                    .=%...:*%%%=.              .#+#*.                                               
                    .+%+.-*###%+-.             .#-.-%=.                                             
              ..:=+#%%#*:--%%%#%*=..  ..--=+-...-%:..*+..                                           
             .*%%%%%#:...=%=%%*%%#*..+*%%%%%%%+..:%=..#+.                                ..         
            .:#%%%+:..  .=+#%#%%%%%=*%%%%%%%%%%%#..*#::#.                      ... :==:.:==:..      
             :+%%%%*..  .++%###+#%#*+%%%%%%%%%%%%=. .:+:.                     :===.-===..:==-:      
        ...  .=%%%%%%#==::*####+=#%#-%%%+#%%%*%%%*.                       .-==:=+=:.....=:....      
      ..#%+. .--%%%%%#*#%#==#%%+*#*-+%#*%%%%%%#*#=.        ..             .===:....:-..  .---:.     
    -%%%%%%+. ..:*%*+%%%%%%%#+=*%+*%%%%%%%%%%#:..::.::.:::#%-.          .===:..--..  .-:..:--:      
    ..  .+%%%+..-#%%%%%%%%%%%%%%**+#%%*++%%%*+...        =%.            .-==:.. ..-..  .-:....      
          +%%%*%%%%%%%%%%%%%%%%%%%#%+#%:.::..            -%+.          .===. .-..  .=..  .---:      
          :%%%%%%%%%%%%%-*%%%%%%%%%-#%#+=.              .-%*           .===.. ..=.   .--...:..      
          .:#%%%%#*==:....:-==-:....=%%%%*=..         ..-#%-            ....-.   .-:.  .---:.       
            ........              .. .:+#%%#=-:::=++**#%%+..           :===...-.   .-:...--..       
                                  ...  ..:-+*%***:.......              .---.  ..-....:---..         
                                  .-.       .++*:-:....                 .:..-....:---.::..          
                                  .:.       .%-    .+=.          ......---=.:---..--:.              
                                  .-.       .%+.   ....     .:--.:---..---:..::.. ..                
                                   :.       .%#.         :--.---...:.:-:.-:.                        
                                   .:       -%*-.      ..---:....:-.   .===:.                       
                                   .:......=%%-.      .---...-:.  ..-:. .::.                        
                                   .+%##%%%#-..      ..:-::.. .:=..  .:--=:.                        
                                  .:*..              :---. .=.. ..-:.. :===.                        
                                                     .:-::  ..=.  ..:-.....                         
                                                     .--..-:  ..-:   .-==-.                         
                                                     :---. .-.. ..-..::--..                         
                                                     ...::. ..=....-===.                            
                                                     .---.=.....===.:-..                            
                                                     .==-..===..+==.                                
                                           .--..===:.-===..=+=.....                                 
                                       .--:===-.-+=::-+=-. ...                                      
                                    ..:===-.::.:-..  .---.                                          
                                    .===:..:-..  .:-.....                                           
                                   .:--.:..  .-:.   .:---

[For support contact elijah@colossal.com]   
''')
# Key Parameters
Organism = "gray wolf"
# GenBank assembly
Accession = "GCA_905319855.2"
# UCSC Browser related to the specified Accession
Browser = "https://genome.ucsc.edu/h/GCA_905319855.2"
# BLAT link found in UCSC Browser related to the specified Accession
BLAT = "https://genome.ucsc.edu/cgi-bin/hgBlat?hgsid=1861123532_8KvRdTM9RSM1nXELY8RKLnjDaLxR&command=start"
# In Silico PCR tool link found in UCSC Browser related to the specified Accession
PCR = "https://genome.ucsc.edu/cgi-bin/hgPcr?hgsid=1861125786_rHScXnM6vOJdAbx9UX3sioLuSWPu"
# NCBI Primer Blast tool link for target on one template
Primer_Blast = "https://www.ncbi.nlm.nih.gov/tools/primer-blast/index.cgi?"
# Excel file path and the name of the sheet containing the gRNA sequences
excel_file_path = '/Users/qichenyuan/Desktop/Prime Editing Wizard/PE_NGS_List.xlsx'
sheet_name = 'Sheet4'

current_time = datetime.now()
formatted_time = current_time.strftime("a, %d %b %Y %H:%M:%S")
print(f"INFO  @ {formatted_time}:\n\tOrganism is {Organism}\n")
print(f"INFO  @ {formatted_time}:\n\tGenome assembly accession is {Accession}\n")
print(f"INFO  @ {formatted_time}:\n\tUCSC Genome Browser link is {Browser}\n")
print(f"INFO  @ {formatted_time}:\n\tBLAT link is {BLAT}\n")
print(f"INFO  @ {formatted_time}:\n\tIn Silico PCR link is {PCR}\n")
print(f"INFO  @ {formatted_time}:\n\tNCBI Primer-Blast link is {Primer_Blast}\n")
print(f"INFO  @ {formatted_time}:\n\tExcel file path is {excel_file_path}\n")
print(f"INFO  @ {formatted_time}:\n\tsheet name is {sheet_name}\n")

# Set up Chrome options for headless mode
chrome_options = Options()
chrome_options.add_argument('--headless')  # This line enables headless mode
# Create a WebDriver instance with the specified options
driver = webdriver.Chrome(options=chrome_options)
# Open Excel file and find sheet with data1
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook[sheet_name]
current_time = datetime.now()
formatted_time = current_time.strftime("%a, %d %b %Y %H:%M:%S")
print(f"INFO  @ {formatted_time}:\n\tAccess target sequences...\n")

# Find the maximum row with data in column B (gRNA sequence column)
max_row = sheet.max_row

# Initialize a counter for non-empty cells
non_empty_count = 0
# Iterate through the rows in the selected column
for row in range(2, max_row + 1):
    cell_value = sheet['F' + str(row)].value
    # Check if the cell is not empty
    if cell_value is not None:
        non_empty_count += 1

# Iteration started
current_time = datetime.now()
formatted_time = current_time.strftime("%a, %d %b %Y %H:%M:%S")
print(f"INFO  @ {formatted_time}:\n\tIteration started...\n")
for i in range(non_empty_count + 2, max_row + 1):
    try:
        # Input the target sequence into the BLAT tool.
        target_sequence = sheet.cell(row=i, column=2).value
        driver.get(BLAT)
        current_time = datetime.now()
        formatted_time = current_time.strftime("%a, %d %b %Y %H:%M:%S")
        print(f"INFO  @ {formatted_time}:\n\tInput gRNA sequence in BLAT tool...\n")
        user_seq_textarea = WebDriverWait(driver, 300).until(
            ec.presence_of_element_located((By.NAME, "userSeq"))
        )
        user_seq_textarea.clear()
        user_seq_textarea.send_keys(target_sequence)

        # Click on "I'm feeling lucky"
        current_time = datetime.now()
        formatted_time = current_time.strftime("%a, %d %b %Y %H:%M:%S")
        print(f"INFO  @ {formatted_time}:\n\tI'm feeling lucky!\n")
        driver.find_element(By.NAME, "Lucky").click()

        # Wait for the view to be present or visible (adjust the timeout as needed)
        wait = WebDriverWait(driver, 300)
        view_element = wait.until(ec.presence_of_element_located((By.XPATH, '//*[@id="view"]/span')))

        # Create an ActionChains object
        actions = ActionChains(driver)

        # Hover over the "View" element to reveal the submenu
        actions.move_to_element(view_element).perform()

        # Locate and click the "DNA" submenu
        dna_submenu = WebDriverWait(driver, 300).until(
            ec.element_to_be_clickable((By.LINK_TEXT, "DNA"))
        )
        dna_submenu.click()

        # Click on "get DNA"
        current_time = datetime.now()
        formatted_time = current_time.strftime("%a, %d %b %Y %H:%M:%S")
        print(f"INFO  @ {formatted_time}:\n\tGet 300 bp genomic DNA sequence flanking each side of the gRNA...\n")

        # Find and input the value "300" at //*[@id="hgSeq.padding5"]
        upstream = WebDriverWait(driver, 300).until(
            ec.presence_of_element_located((By.XPATH, '//*[@id="hgSeq.padding5"]')))
        upstream.clear()  # Clear any existing value
        upstream.send_keys("300")

        # Find and input the value "300" at //*[@id="hgSeq.padding3"]
        downstream = WebDriverWait(driver, 300).until(
            ec.presence_of_element_located((By.XPATH, '//*[@id="hgSeq.padding3"]')))
        downstream.clear()  # Clear any existing value
        downstream.send_keys("300")

        WebDriverWait(driver, 300).until(
            ec.presence_of_element_located((By.ID, "submit"))
        ).click()

        # Get DNA sequence
        dna_sequence = driver.find_element(By.XPATH, '/html/body/pre')
        dna_sequence_text = dna_sequence.text
        lines = dna_sequence_text.split('\n')
        # Concatenate lines 1 to n to isolate the DNA sequence, excluding line 0 containing '>'
        extracted_sequence = ''.join(lines[1:])
        current_time = datetime.now()
        formatted_time = current_time.strftime("%a, %d %b %Y %H:%M:%S")
        print(f"INFO  @ {formatted_time}:\n\tExtracted DNA full length is {len(extracted_sequence)} bp!\n")

        # Export DNA sequence to Excel
        sheet.cell(row=i, column=3, value=extracted_sequence)

        # Save the updated Excel file
        workbook.save(excel_file_path)

        # Launch the browser and navigate to the NCBI Primer Blast Website
        current_time = datetime.now()
        formatted_time = current_time.strftime("%a, %d %b %Y %H:%M:%S")
        print(f"INFO  @ {formatted_time}:\n\tLaunch NCBI Primer Blast tool...\n")
        driver.get(Primer_Blast)

        # Input the value from (row i, column 3) of an Excel file into xPATH //*[@id="seq"]
        current_time = datetime.now()
        formatted_time = current_time.strftime("%a, %d %b %Y %H:%M:%S")
        print(f"INFO  @ {formatted_time}:\n\tInput values...\n")
        excel_value = sheet.cell(row=i, column=3).value
        seq_input = driver.find_element(By.XPATH, '//*[@id="seq"]')
        seq_input.clear()  # Clear any existing value
        seq_input.send_keys(excel_value)

        primer5_start = 1
        primer5_end = 300
        primer3_start = 301 + len(target_sequence)
        primer3_end = 600 + len(target_sequence)

        # Input '1' into the blank with xPATH //*[@id="PRIMER5_START"]
        primer5_start_input = driver.find_element(By.XPATH, '//*[@id="PRIMER5_START"]')
        primer5_start_input.clear()
        primer5_start_input.send_keys(str(primer5_start))

        # Input '300' into the blank with xPATH //*[@id="PRIMER5_END"]
        primer5_end_input = driver.find_element(By.XPATH, '//*[@id="PRIMER5_END"]')
        primer5_end_input.clear()
        primer5_end_input.send_keys(str(primer5_end))

        # Input '321' into the blank with xPATH //*[@id="PRIMER3_START"]
        primer3_start_input = driver.find_element(By.XPATH, '//*[@id="PRIMER3_START"]')
        primer3_start_input.clear()
        primer3_start_input.send_keys(str(primer3_start))

        # Input '620' into the blank with xPATH //*[@id="PRIMER3_END"]
        primer3_end_input = driver.find_element(By.XPATH, '//*[@id="PRIMER3_END"]')
        primer3_end_input.clear()
        primer3_end_input.send_keys(str(primer3_end))

        # Input '250' into the blank with xPATH //*[@id="PRIMER_PRODUCT_MIN"]
        primer_product_min_input = driver.find_element(By.XPATH, '//*[@id="PRIMER_PRODUCT_MIN"]')
        primer_product_min_input.clear()
        primer_product_min_input.send_keys('250')

        # Input '290' into the blank with xPATH //*[@id="PRIMER_PRODUCT_MAX"]
        primer_product_max_input = driver.find_element(By.XPATH, '//*[@id="PRIMER_PRODUCT_MAX"]')
        primer_product_max_input.clear()
        primer_product_max_input.send_keys('290')

        # Select 'custom' under the database with xPATH //*[@id="PRIMER_SPECIFICITY_DATABASE"]/option[6]
        primer_specificity_database_select = driver.find_element(By.XPATH,
                                                                 '//*[@id="PRIMER_SPECIFICITY_DATABASE"]/option[6]')
        primer_specificity_database_select.click()

        # Input accession with xPATH //*[@id="CUSTOM_DB"]
        custom_db_input = driver.find_element(By.XPATH, '//*[@id="CUSTOM_DB"]')
        custom_db_input.clear()
        custom_db_input.send_keys(Accession)

        # Enter ‘Organism’ into this xPATH //*[@id="ORGANISM"]
        organism_input = driver.find_element(By.XPATH, '//*[@id="ORGANISM"]')
        organism_input.clear()
        organism_input.send_keys(Organism)

        # Click on get primers with XPATH //*[@id="searchForm"]/div[3]/div[1]/input
        current_time = datetime.now()
        formatted_time = current_time.strftime("%a, %d %b %Y %H:%M:%S")
        print(f"INFO  @ {formatted_time}:\n\tGet primers!\n")
        get_primers_button = driver.find_element(By.XPATH, '//*[@id="searchForm"]/div[3]/div[1]/input')
        get_primers_button.click()

        # Wait for the checkbox to be present, timeout after 300 seconds
        seq_checkbox = WebDriverWait(driver, 300).until(
            ec.presence_of_element_located((By.XPATH, '//*[@id="descr"]/tbody/tr[2]/td[1]/label'))
        )
        # Click on the checkbox
        current_time = datetime.now()
        formatted_time = current_time.strftime("%a, %d %b %Y %H:%M:%S")
        print(f"INFO  @ {formatted_time}:\n\tChromosome Confirmed!\n")
        seq_checkbox.click()

        # Click on "Submit" with its XPATH //*[@id="userGuidedForm"]/div/div[1]/input
        submit_button = driver.find_element(By.XPATH, '//*[@id="userGuidedForm"]/div/div[1]/input')
        submit_button.click()
        # Wait until you see Primer pair 1 (XPATH = //*[@id="alignments"]/div[1]/h2), typically the best design
        WebDriverWait(driver, 3000).until(
            ec.visibility_of_element_located((By.XPATH, '//*[@id="alignments"]/div[1]/h2')))
        current_time = datetime.now()
        formatted_time = current_time.strftime("%a, %d %b %Y %H:%M:%S")
        print(f"INFO  @ {formatted_time}:\n\tPrimer pair 1 selected!\n")

        # To make sure gRNA sequence can be found in the amplicon sequence for NGS analysis.

        if target_sequence in extracted_sequence:
            current_time = datetime.now()
            formatted_time = current_time.strftime("%a, %d %b %Y %H:%M:%S")
            print(f"INFO  @ {formatted_time}:\n\tTarget found in extracted DNA sequence!\n")
            # Find DNA sequence of Forward primer with its XPATH as //*[@id="alignments"]/div[1]/table/tbody/tr[2]/td[1]
            forward_primer_xpath = '//*[@id="alignments"]/div[1]/table/tbody/tr[2]/td[1]'
            forward_primer = driver.find_element(By.XPATH, forward_primer_xpath).text

            # Find DNA sequence of Reverse primer with its XPATH as //*[@id="alignments"]/div[1]/table/tbody/tr[3]/td[1]
            reverse_primer_xpath = '//*[@id="alignments"]/div[1]/table/tbody/tr[3]/td[1]'
            reverse_primer = driver.find_element(By.XPATH, reverse_primer_xpath).text

        else:
            # gRNA sequence can be found in the reverse strand of the cleaned_dna_sequence
            # Find DNA sequence of Reverse primer with its XPATH as //*[@id="alignments"]/div[1]/table/tbody/tr[2]/td[1]
            reverse_primer_xpath = '//*[@id="alignments"]/div[1]/table/tbody/tr[2]/td[1]'
            reverse_primer = driver.find_element(By.XPATH, reverse_primer_xpath).text

            # Find DNA sequence of Forward primer with its XPATH as //*[@id="alignments"]/div[1]/table/tbody/tr[3]/td[1]
            forward_primer_xpath = '//*[@id="alignments"]/div[1]/table/tbody/tr[3]/td[1]'
            forward_primer = driver.find_element(By.XPATH, forward_primer_xpath).text
            current_time = datetime.now()
            formatted_time = current_time.strftime("%a, %d %b %Y %H:%M:%S")
            print(f"INFO  @ {formatted_time}:\n\tTarget found in reverse strand of the extracted DNA sequence!\n")

        # Product length of Primer pair 1
        product_length_xpath = '//*[@id="alignments"]/div[1]/table/tbody/tr[5]/td'
        product_length = driver.find_element(By.XPATH, product_length_xpath).text
        current_time = datetime.now()
        formatted_time = current_time.strftime("%a, %d %b %Y %H:%M:%S")
        print(f"INFO  @ {formatted_time}:\n\tProduct length of Primer pair 1 is {product_length} bp!\n")

        # Add adapter sequences to primers
        i7_adapter_sequence = "GTCTCGTGGGCTCGGAGATGTGTATAAGAGACAG"
        i5_adapter_sequence = "TCGTCGGCAGCGTCAGATGTGTATAAGAGACAG"
        sheet.cell(row=i, column=4).value = i7_adapter_sequence + forward_primer
        sheet.cell(row=i, column=5).value = i5_adapter_sequence + reverse_primer
        current_time = datetime.now()
        formatted_time = current_time.strftime("%a, %d %b %Y %H:%M:%S")
        print(
            f"INFO  @ {formatted_time}:\n\ti7 adapter sequence is '{i7_adapter_sequence}', added to forward_primer!\n")
        current_time = datetime.now()
        formatted_time = current_time.strftime("%a, %d %b %Y %H:%M:%S")
        print(
            f"INFO  @ {formatted_time}:\n\ti5 adapter sequence is '{i5_adapter_sequence}', added to reverse_primer!\n")
        workbook.save(excel_file_path)

        # Check if the primer pair could generate desired amplicon using UCSC in silico PCR tool
        current_time = datetime.now()
        formatted_time = current_time.strftime("%a, %d %b %Y %H:%M:%S")
        print(f"INFO  @ {formatted_time}:\n\tGenerating amplicon sequence using UCSC in silico PCR tool...\n")

        # Open UCSC genome browser in silico PCR with genome of interest already selected
        driver.get(PCR)

        # Input Forward Primer at //*[@id="wp_f"]
        forward_primer_input = driver.find_element(By.XPATH, '//*[@id="wp_f"]')
        forward_primer_sequence = forward_primer
        forward_primer_input.send_keys(forward_primer_sequence)

        # Input Reverse primer at //*[@id="wp_r"]
        reverse_primer_input = driver.find_element(By.XPATH, '//*[@id="wp_r"]')
        reverse_primer_sequence = reverse_primer
        reverse_primer_input.send_keys(reverse_primer_sequence)

        # Click on the submit button //*[@id="Submit"]
        submit_button = driver.find_element(By.XPATH, '//*[@id="Submit"]')
        submit_button.click()

        # Get DNA sequence
        dna_sequence = WebDriverWait(driver, 300).until(
            ec.presence_of_element_located(
                (
                    By.XPATH,
                    '//*[@id="firstSection"]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[2]/td[2]/tt/pre'))
        )
        dna_sequence_text = dna_sequence.text
        # Count the occurrences of ">", representing each PCR amplicon
        amplicon_number = dna_sequence_text.count('>')
        # Convert text into a list in which the index of '>' related element is 0
        lines = dna_sequence_text.split('\n')
        # Concatenate lines 1 to n to isolate the DNA sequence, excluding line 0 containing '>'
        concatenated_sequence = ''.join(lines[1:])
        amplicon_length = len(concatenated_sequence)
        # Check if off-target products exist and if amplicon length is correct
        if amplicon_number == 1 and amplicon_length == int(product_length):
            sheet.cell(row=i, column=6).value = concatenated_sequence.upper()
            sheet.cell(row=i, column=7).value = len(concatenated_sequence)
            workbook.save(excel_file_path)
            non_empty_count += 1
            current_time = datetime.now()
            formatted_time = current_time.strftime("%a, %d %b %Y %H:%M:%S")
            print(f"INFO  @ {formatted_time}:\n\tTarget-specific amplification!\n")
            print(f"INFO  @ {formatted_time}:\n\t{non_empty_count}/{max_row - 1} completed!\n")
        else:
            # Non-specific PCR, design failed
            sheet.cell(row=i, column=3).value = "Failed"
            sheet.cell(row=i, column=4).value = "Failed"
            sheet.cell(row=i, column=5).value = "Failed"
            sheet.cell(row=i, column=6).value = "Failed"
            sheet.cell(row=i, column=7).value = "Failed"
            workbook.save(excel_file_path)
            non_empty_count += 1
            current_time = datetime.now()
            formatted_time = current_time.strftime("%a, %d %b %Y %H:%M:%S")
            print(f"INFO  @ {formatted_time}:\n\tNon-specific amplification, design failed!\n")
            print(f"INFO  @ {formatted_time}:\n\t{non_empty_count}/{max_row - 1} completed!\n")

    except TimeoutException:
        non_empty_count += 1
        current_time = datetime.now()
        formatted_time = current_time.strftime("%a, %d %b %Y %H:%M:%S")
        print(f"INFO  @ {formatted_time}:\n\tTimeout Exception!\n")
        current_time = datetime.now()
        formatted_time = current_time.strftime("%a, %d %b %Y %H:%M:%S")
        print(f"INFO  @ {formatted_time}:\n\t{non_empty_count}/{max_row - 1} completed!\n")
        # Your code for the next iteration or any other handling
        sheet.cell(row=i, column=3).value = "Failed"
        sheet.cell(row=i, column=4).value = "Failed"
        sheet.cell(row=i, column=5).value = "Failed"
        sheet.cell(row=i, column=6).value = "Failed"
        sheet.cell(row=i, column=7).value = "Failed"
        workbook.save(excel_file_path)
    print('''
                             .=+#%+.                                                                    
                     ..  .-%%%#:                                                                    
                    .=%...:*%%%=.              .#+#*.                                               
                    .+%+.-*###%+-.             .#-.-%=.                                             
              ..:=+#%%#*:--%%%#%*=..  ..--=+-...-%:..*+..                                           
             .*%%%%%#:...=%=%%*%%#*..+*%%%%%%%+..:%=..#+.                                ..         
            .:#%%%+:..  .=+#%#%%%%%=*%%%%%%%%%%%#..*#::#.                      ... :==:.:==:..      
             :+%%%%*..  .++%###+#%#*+%%%%%%%%%%%%=. .:+:.                     :===.-===..:==-:      
        ...  .=%%%%%%#==::*####+=#%#-%%%+#%%%*%%%*.                       .-==:=+=:.....=:....      
      ..#%+. .--%%%%%#*#%#==#%%+*#*-+%#*%%%%%%#*#=.        ..             .===:....:-..  .---:.     
    -%%%%%%+. ..:*%*+%%%%%%%#+=*%+*%%%%%%%%%%#:..::.::.:::#%-.          .===:..--..  .-:..:--:      
    ..  .+%%%+..-#%%%%%%%%%%%%%%**+#%%*++%%%*+...        =%.            .-==:.. ..-..  .-:....      
          +%%%*%%%%%%%%%%%%%%%%%%%#%+#%:.::..            -%+.          .===. .-..  .=..  .---:      
          :%%%%%%%%%%%%%-*%%%%%%%%%-#%#+=.              .-%*           .===.. ..=.   .--...:..      
          .:#%%%%#*==:....:-==-:....=%%%%*=..         ..-#%-            ....-.   .-:.  .---:.       
            ........              .. .:+#%%#=-:::=++**#%%+..           :===...-.   .-:...--..       
                                  ...  ..:-+*%***:.......              .---.  ..-....:---..         
                                  .-.       .++*:-:....                 .:..-....:---.::..          
                                  .:.       .%-    .+=.          ......---=.:---..--:.              
                                  .-.       .%+.   ....     .:--.:---..---:..::.. ..                
                                   :.       .%#.         :--.---...:.:-:.-:.                        
                                   .:       -%*-.      ..---:....:-.   .===:.                       
                                   .:......=%%-.      .---...-:.  ..-:. .::.                        
                                   .+%##%%%#-..      ..:-::.. .:=..  .:--=:.                        
                                  .:*..              :---. .=.. ..-:.. :===.                        
                                                     .:-::  ..=.  ..:-.....                         
                                                     .--..-:  ..-:   .-==-.                         
                                                     :---. .-.. ..-..::--..                         
                                                     ...::. ..=....-===.                            
                                                     .---.=.....===.:-..                            
                                                     .==-..===..+==.                                
                                           .--..===:.-===..=+=.....                                 
                                       .--:===-.-+=::-+=-. ...                                      
                                    ..:===-.::.:-..  .---.                                          
                                    .===:..:-..  .:-.....                                           
                                   .:--.:..  .-:.   .:---
    ''')

# Close the browser
driver.quit()
current_time = datetime.now()
formatted_time = current_time.strftime("%a, %d %b %Y %H:%M:%S")
print(f"INFO  @ {formatted_time}:\n\tANGEL Task Complete!\n")
